import io

import openpyxl
# import xlsxwriter
from langchain.prompts import ChatPromptTemplate
from langchain_openai import ChatOpenAI
import streamlit as st
import pandas as pd
from io import BytesIO

from openpyxl.utils import get_column_letter

# 自定义CSS样式
st.markdown("""
<style>
    .stApp {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        min-height: 100vh;
    }

    .css-1d391kg {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }

    .css-1lcbmhc {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }

    section[data-testid="stSidebar"] {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
    }

    section[data-testid="stSidebar"] label {
        color: white !important;
    }

    section[data-testid="stSidebar"] p {
        color: white !important;
    }

    section[data-testid="stSidebar"] .stTextInput > div > div > label {
        color: white !important;
    }

    .main-header {
        background: linear-gradient(90deg, #667eea 0%, #764ba2 100%);
        padding: 2rem;
        border-radius: 15px;
        margin-bottom: 2rem;
        text-align: center;
        color: white;
        box-shadow: 0 8px 32px rgba(0,0,0,0.2);
    }

    .main-header h1 {
        font-size: 3rem;
        font-weight: bold;
        margin: 0;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }

    .main-header p {
        font-size: 1.2rem;
        margin: 0.5rem 0 0 0;
        opacity: 0.9;
    }

    .content-container {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 8px 32px rgba(0,0,0,0.1);
        margin-bottom: 2rem;
        border: 1px solid #d1d8e0;
    }

    .stButton > button {
        background: linear-gradient(45deg, #667eea, #764ba2);
        color: white;
        border: none;
        padding: 1rem 2rem;
        border-radius: 50px;
        font-size: 1.1rem;
        font-weight: bold;
        cursor: pointer;
        transition: all 0.3s ease;
        box-shadow: 0 4px 15px rgba(102, 126, 234, 0.4);
        width: 100%;
        margin-top: 1rem;
    }

    .stButton > button:hover {
        transform: translateY(-2px);
        box-shadow: 0 6px 20px rgba(102, 126, 234, 0.6);
    }

    .result-container {
        background: linear-gradient(135deg, #f5f7fa 0%, #c3cfe2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        margin-top: 1rem;
        border: 1px solid #e0e0e0;
    }

    .result-title {
        background: linear-gradient(45deg, #ff6b6b, #ee5a24);
        color: white;
        padding: 0.75rem 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        font-weight: bold;
        font-size: 1.1rem;
    }

    .result-content {
        background: white;
        padding: 1rem;
        border-radius: 10px;
        border-left: 4px solid #667eea;
        box-shadow: 0 2px 10px rgba(0,0,0,0.05);
    }

    .sidebar-container {
        background: linear-gradient(180deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 15px;
        color: white;
        margin-bottom: 1rem;
        border: 1px solid #5a6fd8;
    }

    .sidebar-container h3 {
        color: white;
        margin-bottom: 1rem;
        text-align: center;
    }

    .stTextInput > div > div > input {
        border-radius: 10px;
        border: 2px solid #e0e0e0;
        padding: 0.75rem;
        font-size: 1rem;
        background: rgba(255, 255, 255, 0.9);
    }

    .stTextInput > div > div > input:focus {
        border-color: #667eea;
        box-shadow: 0 0 0 0.2rem rgba(102, 126, 234, 0.25);
    }

    .stNumberInput > div > div > input {
        border-radius: 10px;
        border: 2px solid #e0e0e0;
        padding: 0.75rem;
        font-size: 1rem;
        background: rgba(255, 255, 255, 0.9);
    }

    .stSlider > div > div > div > div {
        background: linear-gradient(90deg, #667eea, #764ba2);
    }

    .stSelectbox > div > div > div {
        border-radius: 10px;
        border: 2px solid #e0e0e0;
        background: rgba(255, 255, 255, 0.9);
    }

    .stSelectbox > div > div > div > div {
        background: rgba(255, 255, 255, 0.9);
    }

    .info-box {
        background: linear-gradient(45deg, #74b9ff, #0984e3);
        color: white;
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        text-align: center;
        font-weight: bold;
    }

    .success-box {
        background: linear-gradient(45deg, #00b894, #00a085);
        color: white;
        padding: 1rem 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        text-align: center;
        font-weight: bold;
        font-size: 1.1rem;
    }

    .chat-container {
        background: white;
        border-radius: 15px;
        padding: 1.5rem;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        margin-bottom: 1rem;
    }

    .chat-message {
        background: #f8f9fa;
        padding: 1rem;
        border-radius: 10px;
        margin-bottom: 1rem;
        border-left: 4px solid #667eea;
    }

    .chat-message.user {
        background: #e3f2fd;
        border-left-color: #2196f3;
    }

    .chat-message.ai {
        background: #f3e5f5;
        border-left-color: #9c27b0;
    }

    .file-upload-container {
        background: white;
        padding: 2rem;
        border-radius: 15px;
        box-shadow: 0 4px 20px rgba(0,0,0,0.08);
        text-align: center;
        border: 2px dashed #667eea;
        margin-bottom: 1rem;
    }

</style>
""", unsafe_allow_html=True)

def generate_script(subject,video_length,creativity,api_key):
    title_template=ChatPromptTemplate.from_messages([
        ('human','请为"{subject}"想一个标题')
    ])
    script_template=ChatPromptTemplate.from_messages([
        ('human','''你是一位短视频频道的博主。根据以下标题和相关信息，为短视频频道写一个视频脚本。
        视频标题：{title}，视频时长：{duration}分钟，生成的脚本的长度尽量遵循视频时长的要求。
        要求开头抓住眼球，中间提供干货内容，结尾有惊喜，脚本格式也请按照【开头、中间，结尾】分隔。
        整体内容的表达方式要尽量轻松有趣，吸引年轻人。   ''')
    ])
    model=ChatOpenAI(
        base_url='https://api.openai-hk.com/v1/',
        openai_api_key='hk-cgdzxa100005687542fd4a4dbbe05ceb4fbf98c8be585474',
        temperature=creativity
    )

    #模版->模型
    title_chain = title_template | model
    # 模版->模型
    script_chain = script_template | model

    title = title_chain.invoke({'subject':subject}).content

    script = script_chain.invoke({'title':title,'duration':video_length}).content

    return title,script


def download_script_as_excel2(messages):

    # 将每段消息放入一列中
    df = pd.DataFrame(messages)

    # 创建一个 BytesIO 缓冲区
    buffer = io.BytesIO()

    # 写入 Excel，使用 openpyxl 引擎
    with pd.ExcelWriter(buffer, engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Script')

        # 获取 workbook 和 worksheet 对象
        workbook = writer.book
        worksheet = writer.sheets['Script']

        # 设置所有列宽为 8（包括 index）
        for col_num, column_cells in enumerate(worksheet.columns, 1):  # 从第1列开始
            col_letter = get_column_letter(col_num)
            col_name = df.columns[col_num - 1]
            if 'time' in str(col_name).lower():  # 忽略大小写匹配
                worksheet.column_dimensions[col_letter].width = 22
            else:
                worksheet.column_dimensions[col_letter].width = 16

    buffer.seek(0)

    # 在 Streamlit 中添加下载按钮
    st.download_button(
        label="📥下载脚本为Excel",
        data=buffer,
        file_name="历史视频脚本集合.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True
    )
def download_script_as_excel(messages):

    # 将每段消息放入一列中
    df = pd.DataFrame(messages)

    # 创建一个 BytesIO 缓冲区
    buffer = io.BytesIO()

    # 写入 Excel，使用 openpyxl 引擎
    with pd.ExcelWriter(buffer ,engine='openpyxl') as writer:
        df.to_excel(writer, index=False, sheet_name='Script')

        # 获取 workbook 和 worksheet 对象
        workbook = writer.book
        worksheet = writer.sheets['Script']

        # 设置所有列宽为 8（包括 index）
        for col_num, column_cells in enumerate(worksheet.columns, 1):  # 从第1列开始
            col_letter = get_column_letter(col_num)
            col_name = df.columns[col_num - 1]
            if 'time' in str(col_name).lower():  # 忽略大小写匹配
                worksheet.column_dimensions[col_letter].width = 22
            else:
                worksheet.column_dimensions[col_letter].width = 16

    buffer.seek(0)

    return buffer

